import streamlit as st
import pandas as pd
import numpy as np
import warnings
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import io
import tempfile
import os

warnings.simplefilter(action='ignore', category=FutureWarning)

st.set_page_config(page_title="Procesador 360", layout="wide", page_icon="logo.png", initial_sidebar_state="expanded")

hide_st_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)

def process_self_evaluation(file_obj, sheet_name):
    try:
        file_obj.seek(0)
        df = pd.read_excel(file_obj, sheet_name=sheet_name)
    except Exception as e:
        return pd.DataFrame()

    df.columns = df.columns.str.strip()
    if 'Mi nombre:' not in df.columns:
        return pd.DataFrame()
    df = df.set_index('Mi nombre:')
    cols_to_drop = ['Marca temporal', 'Dirección de correo electrónico']
    existing_cols_to_drop = [col for col in cols_to_drop if col in df.columns]
    df_questions = df.drop(columns=existing_cols_to_drop, errors='ignore')
    mapping = {'Siempre': 4, 'Generalmente': 3, 'Casi nunca': 2, 'Nunca': 1, 'No aplica': np.nan}
    df_numeric = df_questions.replace(mapping)
    question_cols = sorted([col for col in df_numeric.columns if isinstance(col, str) and col.strip().startswith('A.')])
    df_filtered = pd.DataFrame(index=question_cols)
    for col in question_cols:
        if col in df_numeric.columns:
             df_filtered[col] = pd.to_numeric(df_numeric[col], errors='coerce')
    return df_numeric.T.reindex(question_cols).sort_index()

def process_360_detailed_by_role(file_obj, sheet_name):
    try:
        file_obj.seek(0)
        df = pd.read_excel(file_obj, sheet_name=sheet_name)
    except Exception as e:
        return pd.DataFrame()

    df.columns = df.columns.str.strip()
    role_col = 'Posicion'
    name_col = 'Mi nombre:'

    if role_col not in df.columns or name_col not in df.columns:
        return pd.DataFrame()

    mapping = {'Siempre': 4, 'Generalmente': 3, 'Casi nunca': 2, 'Nunca': 1, 'No aplica': np.nan}
    df = df.replace(mapping)
    question_cols = sorted([col for col in df.columns if isinstance(col, str) and col.strip().startswith('A.')])

    for col in question_cols:
        if col in df.columns:
             df[col] = pd.to_numeric(df[col], errors='coerce')

    table_parts = []
    file_obj.seek(0)
    try:
        df_original_for_order = pd.read_excel(file_obj, sheet_name=sheet_name)
        df_original_for_order.columns = df_original_for_order.columns.str.strip()
    except: df_original_for_order = df

    for role in ['Jefe', 'Colega', 'Subordinado']:
        names_in_order = []
        if role_col in df_original_for_order.columns and name_col in df_original_for_order.columns:
            names_in_order = df_original_for_order[df_original_for_order[role_col] == role][name_col].unique()

        role_df = df[df[role_col] == role].copy()
        if role_df.empty: continue

        valid_names_in_order = [name for name in names_in_order if name in role_df[name_col].values]
        missing_names = [name for name in role_df[name_col].unique() if name not in valid_names_in_order]
        final_names_for_role = valid_names_in_order + sorted(missing_names)

        if not final_names_for_role: continue
        valid_question_cols = [q for q in question_cols if q in role_df.columns]
        if not valid_question_cols: continue

        try:
             individual_responses_ordered = role_df.set_index(name_col)[valid_question_cols].reindex(final_names_for_role).T.sort_index()
        except KeyError:
             available_names = [name for name in final_names_for_role if name in role_df[name_col].values]
             if not available_names: continue
             individual_responses_ordered = role_df.set_index(name_col).reindex(available_names)[valid_question_cols].T.sort_index()

        table_parts.append(individual_responses_ordered)
        role_average_df = pd.DataFrame(individual_responses_ordered.mean(axis=1), columns=[f'Promedio {role}']).sort_index()
        table_parts.append(role_average_df)

    if not table_parts:
        return pd.DataFrame(index=question_cols)

    final_table = pd.concat(table_parts, axis=1, join='outer', sort=False)
    return final_table.reindex(question_cols).sort_index()

def format_self_sheet(df):
    if df.empty: return df
    df = df.sort_index()
    individual_cols = df.columns.tolist()
    df['Promedio General'] = np.nan
    for i in range(1, 6):
        section_prefix = f'A.{i}'
        section_questions = [q for q in df.index if isinstance(q, str) and q.strip().startswith(section_prefix)]
        if not section_questions: continue
        last_q = max(section_questions)
        valid_individual_cols = [c for c in individual_cols if c in df.columns]
        if valid_individual_cols and not df.loc[section_questions, valid_individual_cols].isnull().all().all():
             avg = df.loc[section_questions, valid_individual_cols].mean().mean()
             df.loc[last_q, 'Promedio General'] = avg
    return df[[col for col in df.columns if col != 'Promedio General'] + ['Promedio General']]

def format_360_sheet(df, file_obj, sheet_name_360_for_order):
    if df.empty: return df
    df = df.sort_index()
    
    cols_to_process = [col for col in df.columns if col.startswith('Promedio ')]
    new_section_avg_cols = {}
    for avg_col in cols_to_process:
         role_name = avg_col.split(" ")[-1]
         section_averages = pd.Series(index=df.index, dtype=float)
         for i in range(1, 6):
             section_questions = [q for q in df.index if str(q).startswith(f'A.{i}')]
             if not section_questions: continue
             last_q = max(section_questions)
             if avg_col in df.columns:
                  section_averages.loc[last_q] = df.loc[section_questions, avg_col].mean()
         new_section_avg_cols[f'Promedio Sección {role_name}'] = section_averages

    for col_name, col_data in new_section_avg_cols.items(): df[col_name] = col_data

    individual_cols = [col for col in df.columns if not col.startswith('Promedio') and 'Sección' not in col]
    df['Promedio General'] = np.nan
    for i in range(1, 6):
        section_questions = [q for q in df.index if str(q).startswith(f'A.{i}')]
        if not section_questions: continue
        last_q = max(section_questions)
        if individual_cols:
             avg = df.loc[section_questions, individual_cols].mean(skipna=True).mean(skipna=True)
             df.loc[last_q, 'Promedio General'] = avg

    final_order = []
    file_obj.seek(0)
    try:
        df_360_original = pd.read_excel(file_obj, sheet_name=sheet_name_360_for_order)
        df_360_original.columns = df_360_original.columns.str.strip()
    except: df_360_original = None

    role_col, name_col = 'Posicion', 'Mi nombre:'
    for role in ['Jefe', 'Colega', 'Subordinado']:
        if df_360_original is not None and role_col in df_360_original.columns:
            evaluators = df_360_original[df_360_original[role_col] == role][name_col].drop_duplicates().tolist()
            final_order.extend([e for e in evaluators if e in df.columns])
        
        if f'Promedio {role}' in df.columns:
            final_order.append(f'Promedio {role}')
            if f'Promedio Sección {role}' in df.columns:
                final_order.append(f'Promedio Sección {role}')
    
    final_order.append('Promedio General')
    final_cols = [c for c in final_order if c in df.columns]
    missed = [c for c in df.columns if c not in final_cols and c != 'Promedio General']
    pg_idx = final_cols.index('Promedio General') if 'Promedio General' in final_cols else len(final_cols)
    return df[final_cols[:pg_idx] + sorted(missed) + final_cols[pg_idx:]]

def style_worksheet(worksheet, headers):
    if worksheet is None: return
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    avg_blue = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")

    for row in range(2, worksheet.max_row + 1):
        cell = worksheet[f'A{row}']
        if cell: cell.fill = yellow

    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx + 1)
        fill = None
        if header == 'SELF': fill = grey
        elif str(header).startswith('Promedio') and 'Sección' not in str(header) and 'General' not in str(header): fill = blue
        elif 'Sección' in str(header): fill = green
        elif header == 'Promedio General': fill = avg_blue
        
        if fill:
            for row in range(1, worksheet.max_row + 1):
                worksheet[f'{col_letter}{row}'].fill = fill

with st.sidebar:
    st.markdown("### Instrucciones")
    st.info("Sigue estos pasos para generar el reporte:")
    st.markdown(
        """
        1. **Sube el archivo Excel** (.xlsx).
           *Debe contener las pestañas 'SELF ' y '360'.*
        2. **Nombra el reporte** en el cuadro de texto.
        3. Presiona **Generar Reporte**.
        4. Espera a que termine y **descarga**.
        """
    )
    st.divider()
    st.markdown("### Herramientas")
    
    try:
        with open("plantilla_base.xlsx", "rb") as file:
            st.download_button(
                label="📄 Descargar Plantilla Vacía",
                data=file,
                file_name="Plantilla_Evaluacion_360.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except FileNotFoundError:
        st.warning("Archivo 'plantilla_base.xlsx' no encontrado.")

    st.divider()
    st.caption("© 2026 CR Consulting Group")

col1, col2 = st.columns([1, 8], vertical_alignment="center")

with col1:
    st.image("logo.png", width=90)

with col2:
    st.markdown("<h1 style='margin: 0; padding: 0;'>CR Consulting Group Reportes 360</h1>", unsafe_allow_html=True)

st.write("") 

uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"], label_visibility="collapsed")
output_name = st.text_input("Nombre para el reporte final (sin extensión):", "Resultado_Evaluacion_360")

sheet_self = "SELF "
sheet_360 = "360"

if uploaded_file is not None:
    if st.button("Generar Reporte", type="primary"):
        with st.spinner('Procesando datos... esto tomará unos segundos.'):
            try:
                output_buffer = io.BytesIO()
                
                df_self = format_self_sheet(process_self_evaluation(uploaded_file, sheet_self))
                df_360 = format_360_sheet(process_360_detailed_by_role(uploaded_file, sheet_360), uploaded_file, sheet_360)
                
                with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                    if not df_self.empty:
                        df_self.round(1).to_excel(writer, sheet_name='SELF')
                        style_worksheet(writer.sheets['SELF'], df_self.columns)
                    
                    if not df_360.empty:
                        df_360.round(1).to_excel(writer, sheet_name='360')
                        style_worksheet(writer.sheets['360'], df_360.columns)

                output_buffer.seek(0)
                wb = load_workbook(output_buffer)
                
                if not df_self.empty and not df_360.empty:
                    self_avgs = df_self['Promedio General'].dropna().rename('SELF')
                    cols_360 = ['Promedio Sección Jefe', 'Promedio Sección Colega', 'Promedio Sección Subordinado', 'Promedio General']
                    plot_360 = df_360[[c for c in cols_360 if c in df_360.columns]].dropna(how='all')
                    
                    self_avgs = self_avgs.reindex(plot_360.index)
                    combined = pd.concat([self_avgs, plot_360], axis=1)
                    combined.columns = [c.replace('Promedio Sección ', '').replace('Promedio ', '') for c in combined.columns]
                    combined.index = combined.index.str.extract(r'(A\.\d)')[0]
                    combined = combined.groupby(level=0).first().dropna(how='all')

                    if not combined.empty:
                        ax = combined.plot(kind='bar', figsize=(10, 6), rot=0)
                        ax.set_title('Comparativa 360')
                        ax.set_ylim(0, 4.5)
                        plt.tight_layout()
                        
                        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                            plt.savefig(tmp_img.name)
                            tmp_img_path = tmp_img.name
                        
                        if "Gráfica 360" in wb.sheetnames: del wb["Gráfica 360"]
                        ws_chart = wb.create_sheet("Gráfica 360")
                        img = Image(tmp_img_path)
                        ws_chart.add_image(img, 'A1')
                        
                        final_buffer = io.BytesIO()
                        wb.save(final_buffer)
                        output_buffer = final_buffer
                        
                        plt.close()
                        os.unlink(tmp_img_path)

                st.success("¡Archivo procesado correctamente!")
                st.balloons()
                
                final_filename = f"{output_name}.xlsx" if not output_name.endswith('.xlsx') else output_name
                
                st.download_button(
                    label=f"⬇️ Descargar {final_filename}",
                    data=output_buffer.getvalue(),
                    file_name=final_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"Ocurrió un error: {e}")